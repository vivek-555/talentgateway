from openpyxl import load_workbook
import rarfile
import json
import datetime
import boto
import boto.s3
import sys
from boto.s3.key import Key
import os
from django.conf import settings

LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'


def excel_style(col, row=None):
    """ Convert given row and column number to an Excel-style cell name. """
    result = []
    while col:
        col, rem = divmod(col - 1, 26)
        result[:0] = LETTERS[rem]
    if row:
        return ''.join(result) + str(row)
    return ''.join(result)


def read_excel():
    # wb = load_workbook(
    #         filename='~/inception/Sample data upload sheet - Inception.xlsx',
    #         read_only=True)
    #
    # worksheet = wb.worksheets[0]
    #
    # records = []
    #
    # header_row = ['Name', 'Email', 'Gender', 'Secondary Email', 'Date of Birth', 'Mobile',
    #               'UG degree', 'UG completion year', 'UG College', 'UG course', 'PG degree',
    #               'PG completion year', 'PG College', 'PG course', 'Additional degree',
    #               'Additional completion year', 'Additional College', 'Additional course',
    #               'Total Exp', 'Industry', 'Functional Area', 'Primary Skills', 'Secondary Skills',
    #               'Current Company', 'Current Location - City', 'Current Location - Country',
    #               'Current Salary', 'Current Designation', 'Current Exp',
    #               'Current Functional Area', 'Notice period', 'Preferred location',
    #               'Preferred salary', 'Preferred Functional area', 'Preferred industry',
    #               'Email Subject', 'Email Body', 'Applied To', 'Application Date', 'Source',
    #               'Application Category', 'RESUME file Name']
    #
    # header_dict = {}
    # for i in range(1, len(header_row) + 1):
    #     header_dict[header_row[i-1]] = excel_style(i)
    #
    # for i in range(2, worksheet.max_row + 1):
    #     name = worksheet[header_dict["Name"] + str(i)].value
    #     first_name = " ".join(name.split(' ')[:-1] if len(name.split(' ')) > 1 else name)
    #     last_name = " ".join(name.split(' ')[-1:] if len(name.split(' ')) > 1 else "")
    #
    #     email = worksheet[header_dict["Email"] + str(i)].value
    #     username = email
    #
    #     gender = worksheet[header_dict["Gender"] + str(i)].value
    #     gender = "M" if gender.lower() == "male" else "F" if gender.lower() == "female" else "-"
    #
    #     secondary_email = worksheet[header_dict["Secondary Email"] + str(i)].value
    #
    #     dob = worksheet[header_dict["Date of Birth"] + str(i)].value
    #     try:
    #         dob = dob.date()
    #     except ValueError as e:
    #         dob = None
    #
    #     mobile = worksheet[header_dict["Mobile"] + str(i)].value
    #
    #     ug_degree = worksheet[header_dict["UG degree"] + str(i)].value
    #     ug_completion_year = worksheet[header_dict["UG completion year"] + str(i)].value
    #     ug_completion_year = int(ug_completion_year) if ug_completion_year else None
    #     ug_college = worksheet[header_dict["UG College"] + str(i)].value
    #     ug_course = worksheet[header_dict["UG course"] + str(i)].value
    #
    #     pg_degree = worksheet[header_dict["PG degree"] + str(i)].value
    #     pg_completion_year = worksheet[header_dict["PG completion year"] + str(i)].value
    #     pg_completion_year = int(pg_completion_year) if pg_completion_year else None
    #     pg_college = worksheet[header_dict["PG College"] + str(i)].value
    #     pg_course = worksheet[header_dict["PG course"] + str(i)].value
    #
    #     additional_degree = worksheet[header_dict["Additional degree"] + str(i)].value
    #     additional_completion_year = worksheet[header_dict["Additional completion year"] + str(i)].value
    #     additional_completion_year = int(additional_completion_year) if additional_completion_year else None
    #     additional_college = worksheet[header_dict["Additional College"] + str(i)].value
    #     additional_course = worksheet[header_dict["Additional course"] + str(i)].value
    #
    #     total_exp = worksheet[header_dict["Total Exp"] + str(i)].value
    #     total_exp = float(total_exp) if total_exp else None
    #
    #     industry = worksheet[header_dict["Industry"] + str(i)].value
    #     functional_area = worksheet[header_dict["Functional Area"] + str(i)].value
    #
    #     primary_skills = worksheet[header_dict["Primary Skills"] + str(i)].value
    #     secondary_skills = worksheet[header_dict["Secondary Skills"] + str(i)].value
    #
    #     current_company = worksheet[header_dict["Current Company"] + str(i)].value
    #     current_location_city = worksheet[header_dict["Current Location - City"] + str(i)].value
    #     current_location_country = worksheet[header_dict["Current Location - Country"] + str(i)].value
    #
    #     current_salary = worksheet[header_dict["Current Salary"] + str(i)].value
    #     try:
    #         current_salary = int(float(current_salary[4:-7])*100000) if current_salary else None
    #     except ValueError as e:
    #         current_salary = None
    #
    #     current_designation = worksheet[header_dict["Current Designation"] + str(i)].value
    #
    #     current_exp = worksheet[header_dict["Current Exp"] + str(i)].value
    #     try:
    #         current_exp = float(current_exp) if current_exp else None
    #     except ValueError as e:
    #         current_exp = None
    #
    #     current_functional_area = worksheet[header_dict["Current Functional Area"] + str(i)].value
    #     notice_period = worksheet[header_dict["Notice period"] + str(i)].value
    #     preferred_location = worksheet[header_dict["Preferred location"] + str(i)].value
    #
    #     preferred_salary = worksheet[header_dict["Preferred salary"] + str(i)].value
    #     preferred_salary = int(preferred_salary) if preferred_salary else None
    #
    #     preferred_functional_area = worksheet[header_dict["Preferred Functional area"] + str(i)].value
    #     preferred_industry = worksheet[header_dict["Preferred industry"] + str(i)].value
    #     email_subject = worksheet[header_dict["Email Subject"] + str(i)].value
    #     email_body = worksheet[header_dict["Email Body"] + str(i)].value
    #     applied_to = worksheet[header_dict["Applied To"] + str(i)].value
    #     application_date = worksheet[header_dict["Application Date"] + str(i)].value
    #     source = worksheet[header_dict["Source"] + str(i)].value
    #     application_category = worksheet[header_dict["Application Category"] + str(i)].value
    #     resume_file_name = worksheet[header_dict["RESUME file Name"] + str(i)].value
    #
    #     print(locals())
    #
    #
    #     records.append({
    #         "email": email,
    #         "file_name": resume_file_name
    #     })
    AWS_ACCESS_KEY_ID = settings.AWS_SETTINGS["access_key_id"]
    AWS_SECRET_ACCESS_KEY = settings.AWS_SETTINGS["secret_access_key"]
    # bucket_name = 'test-ip-mybucket'
    bucket_name = 'inception-data-bucket'
    # conn = boto.connect_s3(AWS_ACCESS_KEY_ID,
    #                        AWS_SECRET_ACCESS_KEY)

    conn = boto.s3.connect_to_region('ap-south-1',
                                     aws_access_key_id=AWS_ACCESS_KEY_ID,
                                     aws_secret_access_key=AWS_SECRET_ACCESS_KEY
                                     )

    bucketobj = conn.get_bucket(bucket_name)

    rf = rarfile.RarFile('~/inception/Web.rar')
    filename_list = []

    for f in rf.infolist():
        print f.filename, f.file_size
        upload_file = rf.open(f)
        key = boto.s3.key.Key(bucketobj, 'some_file.zip')
        key.send_file(upload_file)

    for filename in rf.namelist()[:-1]:
        filename_list.append(filename.split("/")[-1])

    # for record in records:
    #     if record["file_name"] in filename_list:
    #         record["file_present"] = True

    # print json.dumps(records, indent=4)

    print("done")


def upload_to_s3(aws_access_key_id, aws_secret_access_key, file, bucket, key, callback=None, md5=None,
                 reduced_redundancy=False, content_type=None):
    """
    Uploads the given file to the AWS S3
    bucket and key specified.

    callback is a function of the form:

    def callback(complete, total)

    The callback should accept two integer parameters,
    the first representing the number of bytes that
    have been successfully transmitted to S3 and the
    second representing the size of the to be transmitted
    object.

    Returns boolean indicating success/failure of upload.
    """
    try:
        size = os.fstat(file.fileno()).st_size
    except:
        # Not all file objects implement fileno(),
        # so we fall back on this
        file.seek(0, os.SEEK_END)
        size = file.tell()

    # conn = boto.connect_s3(aws_access_key_id, aws_secret_access_key)

    conn = boto.s3.connect_to_region('ap-south-1',
                                     aws_access_key_id=aws_access_key_id,
                                     aws_secret_access_key=aws_secret_access_key
                                     )

    bucket = conn.get_bucket(bucket, validate=True)
    k = Key(bucket)
    k.key = key
    if content_type:
        k.set_metadata('Content-Type', content_type)
    sent = k.set_contents_from_file(file, cb=callback, md5=md5, reduced_redundancy=reduced_redundancy, rewind=True)

    # Rewind for later use
    file.seek(0)

    url = conn.generate_url(
        6000,
        'GET',
        'inception-data-bucket',
        aws_access_key_id,
        response_headers={
            'response-content-type': 'application/octet-stream'
        })

    if sent == size:
        return True
    return False


def main():
    AWS_ACCESS_KEY_ID = settings.AWS_SETTINGS["access_key_id"]
    AWS_SECRET_ACCESS_KEY = settings.AWS_SETTINGS["secret_access_key"]
    bucket_name = 'inception-data-bucket'

    rf = rarfile.RarFile('~/inception/Profiles.rar')
    filename_list = []

    for f in rf.infolist():
        print f.filename, f.file_size
        upload_file = rf.open(f)

        key = f.filename

        if upload_to_s3(AWS_ACCESS_KEY_ID, AWS_SECRET_ACCESS_KEY, upload_file, bucket_name, key):
            print 'It worked!'
        else:
            print 'The upload failed...'


if __name__ == '__main__':
    main()
