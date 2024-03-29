# Create your tasks here
from __future__ import absolute_import, unicode_literals
from celery import shared_task
from openpyxl import load_workbook
import rarfile
import os
from talent_portal.models import *
from django.utils.crypto import get_random_string
from django.conf import settings
from talent_portal.utils import s3_utils

rarfile.UNRAR_TOOL = "unrar"

base_path = os.path.dirname(os.path.abspath(__file__))

AWS_ACCESS_KEY_ID = settings.AWS_SETTINGS["access_key_id"]
AWS_SECRET_ACCESS_KEY = settings.AWS_SETTINGS["secret_access_key"]
AWS_BUCKET_NAME = settings.AWS_SETTINGS["bucket_name"]
SUPERADMIN_EMAIL_ADDRESS = "admin@abc.com"  # FIXME: change to real email address later


@shared_task
def parse_file(excel_file_name, rar_file_name, batch_id):
    print("Job started")

    excel_status = parse_excel_file(excel_file_name, batch_id)
    rar_status = parse_rar_file(rar_file_name)

    print("excel status: " + str(excel_status["status"]))
    print("rar status: " + str(rar_status["status"]))

    if excel_status["status"] and rar_status["status"]:
        return True
    raise Exception("There was some error")


def parse_rar_file(rar_file_name):
    response = {
        "status": False,
        "data": None,
        "message": None
    }

    try:
        rf = rarfile.RarFile(
            os.path.abspath(os.path.join(base_path, os.pardir, "temp/rar", rar_file_name)))

        for f in rf.infolist():
            try:
                upload_file = rf.open(f)
            except Exception as e:
                continue

            key = f.filename

            if s3_utils.upload_to_s3(AWS_ACCESS_KEY_ID, AWS_SECRET_ACCESS_KEY, upload_file, AWS_BUCKET_NAME, key):
                print 'It worked!'
            else:
                print 'The upload failed...'

    except Exception as e:
        msg = "Error opening file: " + str(e)
        print(msg)
        response["message"] = msg
        return response

    response["status"] = True
    response["message"] = "Extracting rar file done"

    return response


header_row = ['Name', 'Email', 'Gender', 'Secondary Email', 'Date of Birth', 'Mobile',
              'UG degree', 'UG completion year', 'UG College', 'UG course', 'PG degree',
              'PG completion year', 'PG College', 'PG course', 'Additional degree',
              'Additional completion year', 'Additional College', 'Additional course',
              'Total Exp', 'Industry', 'Functional Area', 'Primary Skills', 'Secondary Skills',
              'Current Company', 'Current Location - City', 'Current Location - Country',
              'Current Salary', 'Current Designation', 'Current Exp',
              'Current Functional Area', 'Notice period', 'Preferred location',
              'Preferred salary', 'Preferred Functional area', 'Preferred industry',
              'Email Subject', 'Email Body', 'Applied To', 'Application Date', 'Source',
              'Application Category', 'RESUME file Name']


def excel_style(col, row=None):
    """ Convert given row and column number to an Excel-style cell name. """
    LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    result = []
    while col:
        col, rem = divmod(col - 1, 26)
        result[:0] = LETTERS[rem]
    if row:
        return ''.join(result) + str(row)
    return ''.join(result)


header_dict = {}
for i in range(1, len(header_row) + 1):
    header_dict[header_row[i - 1]] = excel_style(i)


def parse_excel_file(excel_file_name, batch_id):
    response = {
        "status": False,
        "data": None,
        "message": None
    }

    try:
        wb = load_workbook(
            filename=os.path.abspath(
                os.path.join(base_path, os.pardir, "temp/excel", excel_file_name)))

    except Exception as e:
        msg = "Error opening file: " + str(e)
        print(msg)
        response["message"] = msg
        return response

    worksheet = wb.worksheets[0]

    for i in range(2, worksheet.max_row + 1):
        email = worksheet[header_dict["Email"] + str(i)].value
        if not email:
            email = "noemailfound-" + get_random_string(length=4) + "@osttalent.com"

        try:
            customer = Customer.objects.get(email=email)
        except Customer.DoesNotExist as e:
            customer = None

        username = email

        name = worksheet[header_dict["Name"] + str(i)].value

        try:
            first_name = " ".join(name.split(' ')[:-1]) if len(name.split(' ')) > 1 else name
            last_name = " ".join(name.split(' ')[-1:]) if len(name.split(' ')) > 1 else ""
        except Exception as e:
            first_name = ""
            last_name = ""

        gender = worksheet[header_dict["Gender"] + str(i)].value
        try:
            gender = "M" if gender.lower() == "male" else "F" if gender.lower() == "female" else "-"
        except Exception as e:
            gender = "-"

        secondary_email = worksheet[header_dict["Secondary Email"] + str(i)].value

        dob = worksheet[header_dict["Date of Birth"] + str(i)].value
        try:
            dob = dob.date()
        except Exception as e:
            dob = None

        mobile = worksheet[header_dict["Mobile"] + str(i)].value

        ug_degree = worksheet[header_dict["UG degree"] + str(i)].value
        ug_completion_year = worksheet[header_dict["UG completion year"] + str(i)].value
        ug_completion_year = int(ug_completion_year) if ug_completion_year else None
        ug_college = worksheet[header_dict["UG College"] + str(i)].value
        ug_course = worksheet[header_dict["UG course"] + str(i)].value

        pg_degree = worksheet[header_dict["PG degree"] + str(i)].value
        pg_completion_year = worksheet[header_dict["PG completion year"] + str(i)].value
        pg_completion_year = int(pg_completion_year) if pg_completion_year else None
        pg_college = worksheet[header_dict["PG College"] + str(i)].value
        pg_course = worksheet[header_dict["PG course"] + str(i)].value

        additional_degree = worksheet[header_dict["Additional degree"] + str(i)].value
        additional_completion_year = worksheet[
            header_dict["Additional completion year"] + str(i)].value
        additional_completion_year = int(
            additional_completion_year) if additional_completion_year else None
        additional_college = worksheet[header_dict["Additional College"] + str(i)].value
        additional_course = worksheet[header_dict["Additional course"] + str(i)].value

        total_exp = worksheet[header_dict["Total Exp"] + str(i)].value
        total_exp = float(total_exp) if total_exp else None

        industry = worksheet[header_dict["Industry"] + str(i)].value
        functional_area = worksheet[header_dict["Functional Area"] + str(i)].value

        primary_skills = worksheet[header_dict["Primary Skills"] + str(i)].value
        secondary_skills = worksheet[header_dict["Secondary Skills"] + str(i)].value

        current_company = worksheet[header_dict["Current Company"] + str(i)].value
        current_location_city = worksheet[header_dict["Current Location - City"] + str(i)].value
        current_location_country = worksheet[
            header_dict["Current Location - Country"] + str(i)].value

        current_salary = worksheet[header_dict["Current Salary"] + str(i)].value
        try:
            current_salary = int(float(current_salary[4:-7]) * 100000) if current_salary else None
        except ValueError as e:
            current_salary = None

        current_designation = worksheet[header_dict["Current Designation"] + str(i)].value

        current_exp = worksheet[header_dict["Current Exp"] + str(i)].value
        try:
            current_exp = float(current_exp) if current_exp else None
        except ValueError as e:
            current_exp = None

        current_functional_area = worksheet[header_dict["Current Functional Area"] + str(i)].value
        notice_period = worksheet[header_dict["Notice period"] + str(i)].value
        preferred_location = worksheet[header_dict["Preferred location"] + str(i)].value

        preferred_salary = worksheet[header_dict["Preferred salary"] + str(i)].value
        preferred_salary = int(preferred_salary) if preferred_salary else None

        preferred_functional_area = worksheet[
            header_dict["Preferred Functional area"] + str(i)].value
        preferred_industry = worksheet[header_dict["Preferred industry"] + str(i)].value
        email_subject = worksheet[header_dict["Email Subject"] + str(i)].value
        email_body = worksheet[header_dict["Email Body"] + str(i)].value
        applied_to = worksheet[header_dict["Applied To"] + str(i)].value
        application_date = worksheet[header_dict["Application Date"] + str(i)].value
        try:
            application_date = application_date.date()
        except Exception as e:
            application_date = None
        source = worksheet[header_dict["Source"] + str(i)].value
        application_category = worksheet[header_dict["Application Category"] + str(i)].value
        resume_file_name = worksheet[header_dict["RESUME file Name"] + str(i)].value

        try:
            resume_text = email_subject + "\n" + email_body
        except Exception as e:
            resume_text = None

        if customer:
            customer.first_name = first_name if first_name else customer.first_name
            customer.last_name = last_name if last_name else customer.last_name

            customer.gender = gender if gender else customer.gender
            customer.secondary_email = secondary_email if secondary_email else customer.secondary_email
            customer.dob = dob if dob else customer.dob
            customer.mobile = mobile if mobile else customer.mobile
            customer.ug_degree = ug_degree if ug_degree else customer.ug_degree
            customer.ug_completion_year = ug_completion_year if ug_completion_year else customer.ug_completion_year
            customer.ug_college = ug_college if ug_college else customer.ug_college
            customer.ug_course = ug_course if ug_course else customer.ug_course
            customer.pg_degree = pg_degree if pg_degree else customer.pg_degree
            customer.pg_completion_year = pg_completion_year if pg_completion_year else customer.pg_completion_year
            customer.pg_college = pg_college if pg_college else customer.pg_college
            customer.pg_course = pg_course if pg_course else customer.pg_course
            customer.additional_degree = additional_degree if additional_degree else customer.additional_degree
            customer.additional_completion_year = additional_completion_year if additional_completion_year else customer.additional_completion_year
            customer.additional_college = additional_college if additional_college else customer.additional_college
            customer.additional_course = additional_course if additional_course else customer.additional_course
            customer.total_experience = total_exp if total_exp else customer.total_experience
            customer.current_company = current_company if current_company else customer.current_company
            customer.current_location = current_location_city if current_location_city else customer.current_location
            customer.current_salary = current_salary if current_salary else customer.current_salary
            customer.current_designation = current_designation if current_designation else customer.current_designation
            customer.current_exprience = current_exp if current_exp else customer.current_exprience
            customer.industry = industry if industry else customer.industry
            customer.primary_skills = primary_skills if primary_skills else customer.primary_skills
            customer.secondary_skills = secondary_skills if secondary_skills else customer.secondary_skills
            customer.current_functional_area = current_functional_area if current_functional_area else customer.current_functional_area
            customer.notice_period = notice_period if notice_period else customer.notice_period
            customer.preferred_location = preferred_location if preferred_location else customer.preferred_location
            customer.preferred_salary = preferred_salary if preferred_salary else customer.preferred_salary
            customer.preferred_functional_area = preferred_functional_area if preferred_functional_area else customer.preferred_functional_area
            customer.preferred_industry = preferred_industry if preferred_industry else customer.preferred_industry
            customer.resume_url = resume_file_name if resume_file_name else customer.resume_url
            customer.resume_text = resume_text if resume_text else customer.resume_text
            customer.created_by = SUPERADMIN_EMAIL_ADDRESS,
            customer.batch_id = batch_id if batch_id else customer.batch_id
            customer.applied_to = applied_to if applied_to else customer.applied_to
            customer.source = source if source else customer.source
            customer.application_date = application_date if application_date else customer.application_date
            customer.category = application_category if application_category else customer.category

            try:
                customer.save()

            except Exception as e:
                msg = "Error while saving customer: " + str(e)
                print(msg)

        else:
            #  if not an existing customer, creating a new record
            try:
                customer = Customer(email=email, first_name=first_name, last_name=last_name,
                                    gender=gender,
                                    secondary_email=secondary_email, dob=dob, mobile=mobile,
                                    ug_degree=ug_degree, ug_completion_year=ug_completion_year,
                                    ug_college=ug_college, ug_course=ug_course,
                                    pg_degree=pg_degree,
                                    pg_completion_year=pg_completion_year, pg_college=pg_college,
                                    pg_course=pg_course, additional_degree=additional_degree,
                                    additional_completion_year=additional_completion_year,
                                    additional_college=additional_college,
                                    additional_course=additional_course,
                                    total_experience=total_exp,
                                    industry=industry,
                                    primary_skills=primary_skills,
                                    secondary_skills=secondary_skills,
                                    current_company=current_company,
                                    current_location=current_location_city,
                                    current_salary=current_salary,
                                    current_designation=current_designation,
                                    current_exprience=current_exp,
                                    current_functional_area=current_functional_area,
                                    notice_period=notice_period,
                                    preferred_location=preferred_location,
                                    preferred_salary=preferred_salary,
                                    preferred_functional_area=preferred_functional_area,
                                    preferred_industry=preferred_industry, resume_text=resume_text,
                                    source=source, application_date=application_date,
                                    applied_to=applied_to, category=application_category,
                                    resume_url=resume_file_name,
                                    created_by=SUPERADMIN_EMAIL_ADDRESS,
                                    batch_id=batch_id, username=username)

                customer.save()
            except Exception as e:
                msg = "Error while saving customer: " + str(e)
                print(msg)

        update_college([pg_college, ug_college, additional_college])
        update_company([current_company])
        update_course([pg_course, ug_course, additional_course])
        update_degree([pg_degree, ug_degree, additional_degree])
        update_functional_area([current_functional_area, preferred_functional_area])
        update_industry([industry, preferred_industry])
        skills = [] if not primary_skills else primary_skills.split(",")
        if secondary_skills:
            skills.extend(secondary_skills.split(","))
        update_skills(skills)

    response["status"] = True
    response["message"] = "Parsing excel file done"

    return response


def update_college(object_list):
    for object in object_list:
        if not object:
            continue
        try:
            if College.objects.filter(name=object):
                continue
            obj = College(name=object, created_by=SUPERADMIN_EMAIL_ADDRESS)
            obj.save()
        except Exception as e:
            print("There was some error " + str(e))


def update_company(object_list):
    for object in object_list:
        if not object:
            continue
        try:
            if Company.objects.filter(name=object):
                continue
            obj = Company(name=object, created_by=SUPERADMIN_EMAIL_ADDRESS)
            obj.save()
        except Exception as e:
            print("There was some error " + str(e))


def update_course(object_list):
    for object in object_list:
        if not object:
            continue
        try:
            if Course.objects.filter(name=object):
                continue
            obj = Course(name=object, created_by=SUPERADMIN_EMAIL_ADDRESS)
            obj.save()
        except Exception as e:
            print("There was some error " + str(e))


def update_degree(object_list):
    for object in object_list:
        if not object:
            continue
        try:
            if Degree.objects.filter(name=object):
                continue
            obj = Degree(name=object, created_by=SUPERADMIN_EMAIL_ADDRESS)
            obj.save()
        except Exception as e:
            print("There was some error " + str(e))


def update_industry(object_list):
    for object in object_list:
        if not object:
            continue
        try:
            if Industry.objects.filter(name=object):
                continue
            obj = Industry(name=object, created_by=SUPERADMIN_EMAIL_ADDRESS)
            obj.save()
        except Exception as e:
            print("There was some error " + str(e))


def update_functional_area(object_list):
    for object in object_list:
        if not object:
            continue
        try:
            if FunctionalArea.objects.filter(name=object):
                continue
            obj = FunctionalArea(name=object, created_by=SUPERADMIN_EMAIL_ADDRESS)
            obj.save()
        except Exception as e:
            print("There was some error " + str(e))


def update_skills(object_list):
    for object in object_list:
        if not object:
            continue
        try:
            if Skill.objects.filter(name=object):
                continue
            obj = Skill(name=object, created_by=SUPERADMIN_EMAIL_ADDRESS)
            obj.save()
        except Exception as e:
            print("There was some error " + str(e))
